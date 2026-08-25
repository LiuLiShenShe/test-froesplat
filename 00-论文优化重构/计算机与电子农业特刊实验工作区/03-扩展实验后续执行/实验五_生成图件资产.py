#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate visual assets and source data for experiment five.

Current reproducible scope:
- Manual-vs-virtual phenotype agreement from ``植株数据.xlsx``.
- Trait-level sensitivity ranking based on MAE/MAPE/R2.
- A status table that keeps the strict multi-mask reconstruction experiment
  separate from the completed phenotype-agreement evidence.

The script does not fabricate multi-mask reconstruction results. Rows that
require SEEM/SAM3/RAP-FSAM3-v2 masks to enter the same 2DGS/mesh/phenotyping
pipeline are written as pending status records.
"""

from __future__ import annotations

import argparse
import csv
import math
import shutil
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import font_manager
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parent
PAPER_ROOT = WORKSPACE_DIR.parent
VIS_ROOT = WORKSPACE_DIR / "05-图件与论文映射/实验五_掩膜到表型误差敏感性"
SOURCE_DATA_ROOT = VIS_ROOT / "source_data"
RESULT_DIR = WORKSPACE_DIR / "04-结果表格模板"

PHENOTYPE_XLSX = PAPER_ROOT / "植株数据.xlsx"
TRAIT_SUMMARY_CSV = RESULT_DIR / "实验五_人工虚拟表型汇总.csv"
EXP5_TABLE_CSV = RESULT_DIR / "实验五_掩膜到表型误差结果表.csv"

TRAITS = ["株高", "冠幅", "叶长", "叶宽"]
TRAIT_COLOR = {
    "株高": "#4c78a8",
    "冠幅": "#59a14f",
    "叶长": "#f28e2b",
    "叶宽": "#e15759",
}


def configure_fonts() -> None:
    candidates = [
        "Noto Sans CJK SC",
        "Noto Sans CJK JP",
        "Source Han Sans SC",
        "WenQuanYi Micro Hei",
        "DejaVu Sans",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.sans-serif"] = [name]
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["pdf.fonttype"] = 42
    plt.rcParams["ps.fonttype"] = 42


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def load_phenotype_pairs() -> list[dict[str, object]]:
    wb = load_workbook(PHENOTYPE_XLSX, data_only=True, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    rows: list[dict[str, object]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        sample = str(row[0]) if row and row[0] is not None else f"plant_{row_index:02d}"
        for trait, truth_col, virtual_col, replicate in [
            ("株高", "株高真值", "株高虚拟植", 1),
            ("冠幅", "冠幅真值", "冠幅虚拟植", 1),
        ]:
            manual = float(row[idx[truth_col]])
            virtual = float(row[idx[virtual_col]])
            rows.append(pair_row(sample, trait, replicate, manual, virtual))

        for replicate in (1, 2, 3):
            for trait, truth_prefix, virtual_prefix in [
                ("叶长", "叶长真值", "叶长虚拟植"),
                ("叶宽", "叶宽真值", "叶宽虚拟植"),
            ]:
                manual = float(row[idx[f"{truth_prefix}{replicate}"]])
                virtual = float(row[idx[f"{virtual_prefix}{replicate}"]])
                rows.append(pair_row(sample, trait, replicate, manual, virtual))
    return rows


def pair_row(sample: str, trait: str, replicate: int, manual: float, virtual: float) -> dict[str, object]:
    error = virtual - manual
    abs_error = abs(error)
    ape = abs_error / manual * 100.0 if manual else math.nan
    return {
        "sample": sample,
        "trait": trait,
        "replicate": replicate,
        "manual": manual,
        "virtual": virtual,
        "error": error,
        "abs_error": abs_error,
        "ape_percent": ape,
    }


def trait_metric_map() -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = {}
    for row in read_csv(TRAIT_SUMMARY_CSV):
        trait = row["性状"]
        out[trait] = {
            "n": float(row["n"]),
            "manual_mean": float(row["人工均值"]),
            "virtual_mean": float(row["虚拟均值"]),
            "mae": float(row["MAE"]),
            "rmse": float(row["RMSE"]),
            "mape": float(row["MAPE百分比"]),
            "bias": float(row["偏差"]),
            "r2": float(row["R2"]),
        }
    return out


def sensitivity_rows(metrics: dict[str, dict[str, float]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    ordered = sorted(TRAITS, key=lambda t: (-metrics[t]["mape"], metrics[t]["r2"]))
    for rank, trait in enumerate(ordered, start=1):
        m = metrics[trait]
        rows.append(
            {
                "rank_by_mape_desc": rank,
                "trait": trait,
                "n": int(m["n"]),
                "MAE": f"{m['mae']:.4f}",
                "RMSE": f"{m['rmse']:.4f}",
                "MAPE_percent": f"{m['mape']:.4f}",
                "Bias": f"{m['bias']:.4f}",
                "R2": f"{m['r2']:.4f}",
                "interpretation": "boundary-sensitive local trait" if trait == "叶宽" else "global or less boundary-local trait",
            }
        )
    return rows


def strict_status_rows() -> list[dict[str, object]]:
    return [
        {
            "evidence": "manual_virtual_phenotype_agreement",
            "status": "completed_current_reproducible",
            "scope": "21 plants / 10 varieties; height, canopy width, three leaf lengths, three leaf widths",
            "usable_claim": "leaf width has the highest MAPE and lowest R2 in the completed pipeline",
            "not_supported_claim": "specific segmentation models causally produce specific phenotype errors",
        },
        {
            "evidence": "SEEM_SAM3_RAP-FSAM3_multimask_reconstruction",
            "status": "pending_new_reconstruction",
            "scope": "same 2DGS/mesh/phenotyping pipeline for each mask source",
            "usable_claim": "",
            "not_supported_claim": "strict mask-source-to-phenotype sensitivity comparison",
        },
        {
            "evidence": "erosion_dilation_control_masks",
            "status": "pending_new_reconstruction",
            "scope": "controlled mask perturbation followed by reconstruction and measurement",
            "usable_claim": "",
            "not_supported_claim": "quantitative correlation between perturbation magnitude and trait error",
        },
    ]


def archive_source_data(pair_rows: list[dict[str, object]], metrics: dict[str, dict[str, float]]) -> list[dict[str, object]]:
    SOURCE_DATA_ROOT.mkdir(parents=True, exist_ok=True)
    asset_rows: list[dict[str, object]] = []
    for src in [TRAIT_SUMMARY_CSV, EXP5_TABLE_CSV]:
        dst = SOURCE_DATA_ROOT / src.name
        shutil.copy2(src, dst)
        asset_rows.append({"asset_type": "source_data", "name": src.name, "path": str(dst)})

    pair_path = SOURCE_DATA_ROOT / "phenotype_pair_errors.csv"
    write_csv(
        pair_path,
        ["sample", "trait", "replicate", "manual", "virtual", "error", "abs_error", "ape_percent"],
        pair_rows,
    )
    asset_rows.append({"asset_type": "source_data", "name": pair_path.name, "path": str(pair_path)})

    sensitivity_path = SOURCE_DATA_ROOT / "trait_sensitivity_ranking.csv"
    write_csv(
        sensitivity_path,
        ["rank_by_mape_desc", "trait", "n", "MAE", "RMSE", "MAPE_percent", "Bias", "R2", "interpretation"],
        sensitivity_rows(metrics),
    )
    asset_rows.append({"asset_type": "source_data", "name": sensitivity_path.name, "path": str(sensitivity_path)})

    status_path = SOURCE_DATA_ROOT / "strict_multimask_status.csv"
    write_csv(
        status_path,
        ["evidence", "status", "scope", "usable_claim", "not_supported_claim"],
        strict_status_rows(),
    )
    asset_rows.append({"asset_type": "source_data", "name": status_path.name, "path": str(status_path)})
    return asset_rows


def save_figure(fig: plt.Figure, path: Path) -> list[Path]:
    path.parent.mkdir(parents=True, exist_ok=True)
    written = [path]
    fig.savefig(path, dpi=600, bbox_inches="tight")
    for suffix in [".pdf", ".svg", ".tif"]:
        out = path.with_suffix(suffix)
        kwargs = {"bbox_inches": "tight"}
        if suffix == ".tif":
            kwargs["dpi"] = 600
        fig.savefig(out, **kwargs)
        written.append(out)
    plt.close(fig)
    return written


def make_trait_profile(metrics: dict[str, dict[str, float]], pair_rows: list[dict[str, object]]) -> list[Path]:
    x = np.arange(len(TRAITS))
    colors = [TRAIT_COLOR[t] for t in TRAITS]
    fig, axes = plt.subplots(2, 2, figsize=(11.8, 8.6))
    fig.suptitle("Experiment 5 current reproducible phenotype sensitivity", fontsize=15, fontweight="bold")

    axes[0, 0].bar(x, [metrics[t]["mae"] for t in TRAITS], color=colors)
    axes[0, 0].set_xticks(x, TRAITS)
    axes[0, 0].set_ylabel("MAE")
    axes[0, 0].set_title("Absolute phenotype error")

    axes[0, 1].bar(x, [metrics[t]["mape"] for t in TRAITS], color=colors)
    axes[0, 1].set_xticks(x, TRAITS)
    axes[0, 1].set_ylabel("MAPE (%)")
    axes[0, 1].set_title("Relative error highlights leaf width")

    axes[1, 0].bar(x, [metrics[t]["r2"] for t in TRAITS], color=colors)
    axes[1, 0].set_xticks(x, TRAITS)
    axes[1, 0].set_ylim(0.84, 1.01)
    axes[1, 0].set_ylabel("R2")
    axes[1, 0].set_title("Agreement with manual measurement")

    ax = axes[1, 1]
    for trait in TRAITS:
        rows = [row for row in pair_rows if row["trait"] == trait]
        ax.scatter(
            [float(row["manual"]) for row in rows],
            [float(row["virtual"]) for row in rows],
            s=28,
            alpha=0.78,
            color=TRAIT_COLOR[trait],
            label=trait,
            edgecolors="white",
            linewidths=0.5,
        )
    all_values = [float(row["manual"]) for row in pair_rows] + [float(row["virtual"]) for row in pair_rows]
    lo, hi = min(all_values), max(all_values)
    pad = (hi - lo) * 0.04
    ax.plot([lo - pad, hi + pad], [lo - pad, hi + pad], color="#555555", linewidth=1.2, linestyle="--")
    ax.set_xlim(lo - pad, hi + pad)
    ax.set_ylim(lo - pad, hi + pad)
    ax.set_xlabel("Manual")
    ax.set_ylabel("Virtual")
    ax.set_title("Manual vs virtual values")
    ax.legend(frameon=False, fontsize=9, ncols=2)

    for ax in axes.ravel():
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(axis="y", color="#d9d9d9", linewidth=0.6, alpha=0.7)

    fig.text(
        0.01,
        0.01,
        "Current claim: completed pipeline supports trait-level sensitivity ranking; strict multi-mask source comparison remains pending.",
        fontsize=9,
        color="#555555",
    )
    return save_figure(fig, VIS_ROOT / "figures/Fig_E5_trait_error_profile.png")


def make_error_distribution(pair_rows: list[dict[str, object]]) -> list[Path]:
    data = [[float(row["ape_percent"]) for row in pair_rows if row["trait"] == trait] for trait in TRAITS]
    fig, ax = plt.subplots(figsize=(9.2, 5.6))
    parts = ax.violinplot(data, showmeans=True, showextrema=False)
    for body, trait in zip(parts["bodies"], TRAITS):
        body.set_facecolor(TRAIT_COLOR[trait])
        body.set_edgecolor("#333333")
        body.set_alpha(0.72)
    parts["cmeans"].set_color("#222222")
    parts["cmeans"].set_linewidth(1.4)
    ax.boxplot(
        data,
        positions=np.arange(1, len(TRAITS) + 1),
        widths=0.16,
        patch_artist=True,
        boxprops={"facecolor": "white", "edgecolor": "#222222", "linewidth": 1.0},
        medianprops={"color": "#222222", "linewidth": 1.2},
        whiskerprops={"color": "#222222", "linewidth": 0.9},
        capprops={"color": "#222222", "linewidth": 0.9},
        flierprops={"marker": "o", "markersize": 3, "markerfacecolor": "#777777", "markeredgecolor": "none", "alpha": 0.45},
    )
    ax.set_xticks(np.arange(1, len(TRAITS) + 1), TRAITS)
    ax.set_ylabel("Absolute percentage error (%)")
    ax.set_title("Trait-level phenotype error distribution")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", color="#d9d9d9", linewidth=0.6, alpha=0.7)
    return save_figure(fig, VIS_ROOT / "figures/Fig_E5_error_distribution.png")


def markdown_metric_table(metrics: dict[str, dict[str, float]]) -> str:
    lines = [
        "| Trait | n | MAE | RMSE | MAPE (%) | Bias | R2 |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for trait in TRAITS:
        m = metrics[trait]
        lines.append(
            f"| {trait} | {int(m['n'])} | {m['mae']:.4f} | {m['rmse']:.4f} | "
            f"{m['mape']:.4f} | {m['bias']:.4f} | {m['r2']:.4f} |"
        )
    return "\n".join(lines)


def write_readme(asset_rows: list[dict[str, object]], metrics: dict[str, dict[str, float]]) -> None:
    readme = f"""# 实验五 掩膜到表型误差敏感性图件

## 存放内容

- `figures/Fig_E5_trait_error_profile.*`：当前可复现人工-虚拟表型一致性图板，含 MAE、MAPE、R2 和人工-虚拟散点。
- `figures/Fig_E5_error_distribution.*`：四类性状绝对百分比误差分布。
- `source_data/phenotype_pair_errors.csv`：从 `植株数据.xlsx` 展开的逐样本/逐性状误差。
- `source_data/trait_sensitivity_ranking.csv`：按 MAPE 降序排序的性状敏感性表。
- `source_data/strict_multimask_status.csv`：严格多掩膜重建敏感性实验的完成边界。
- `figure_asset_index.csv`：本目录全部图件与 source_data 索引。

## 当前核心结果

{markdown_metric_table(metrics)}

## 写作边界

- 当前可写：完整流程中，叶宽的 MAPE 最高且 R2 最低，可作为边界敏感局部性状的证据。
- 不能写成已完成：SEEM、SAM3 单提示词、RAP-FSAM3-v2 或人工腐蚀/膨胀掩膜分别驱动同一 2DGS/网格/表型流程后的因果比较。
- 严格多掩膜敏感性仍需补跑统一重建与虚拟测量后，再统计表型误差与 HD95、边界 F、外部非黑比例、泄漏能量之间的相关性。
"""
    (VIS_ROOT / "图板说明.md").write_text(readme, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate experiment-five visual assets.")
    parser.parse_args()
    configure_fonts()
    pair_rows = load_phenotype_pairs()
    metrics = trait_metric_map()

    asset_rows: list[dict[str, object]] = []
    asset_rows.extend(archive_source_data(pair_rows, metrics))
    for path in make_trait_profile(metrics, pair_rows):
        asset_rows.append({"asset_type": f"trait_profile_{path.suffix.lstrip('.')}", "name": path.name, "path": str(path)})
    for path in make_error_distribution(pair_rows):
        asset_rows.append({"asset_type": f"error_distribution_{path.suffix.lstrip('.')}", "name": path.name, "path": str(path)})

    write_readme(asset_rows, metrics)
    asset_rows.append({"asset_type": "readme", "name": "图板说明.md", "path": str(VIS_ROOT / "图板说明.md")})
    write_csv(VIS_ROOT / "figure_asset_index.csv", ["asset_type", "name", "path"], asset_rows)
    print(f"Wrote {len(asset_rows)} experiment-five assets under {VIS_ROOT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
