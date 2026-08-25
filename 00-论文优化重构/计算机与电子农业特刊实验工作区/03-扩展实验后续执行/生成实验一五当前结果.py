#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate current reproducible tables for experiment 1 and experiment 5.

The script only uses local evidence that already exists in this workspace:
- SAM3/RAP-FSAM3-v2 mask metrics from S21/S22 CSV files.
- Historical SEEM aggregate from the existing evidence note.
- Manual-vs-virtual phenotype workbook.

External baselines that have not been run locally are written as explicit
"pending new inference" rows. X-Decoder and OpenSeeD are now intentionally
paused for the current paper-closing loop, not treated as next-run blockers.
"""

from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parent
PAPER_ROOT = WORKSPACE_DIR.parent
PROJECT_ROOT = PAPER_ROOT.parent
RESULT_DIR = WORKSPACE_DIR / "04-结果表格模板"
DATA_ROOT = PAPER_ROOT / "数据管理"


EXP1_SEGMENTATION_TABLE = RESULT_DIR / "实验一_视觉基础模型横向对比分割表.csv"
EXP1_RECON_TABLE = RESULT_DIR / "实验一_视觉基础模型横向对比下游重建表.csv"
EXP5_TABLE = RESULT_DIR / "实验五_掩膜到表型误差结果表.csv"
EXP5_TRAIT_SUMMARY = RESULT_DIR / "实验五_人工虚拟表型汇总.csv"

KQ_SUMMARY = (
    DATA_ROOT
    / "05-评测结果/S21_KongQueZhuYu_E2_E3/E3v2_KongQueZhuYu_GT5_full_metrics/summary_metrics.csv"
)
XKL_SUMMARY = (
    DATA_ROOT
    / "05-评测结果/S22_XianKeLai1_RAP_FSAM3_GT1/v2_a1s_a5c_metrics/summary_metrics.csv"
)
S23_SUMMARY = (
    DATA_ROOT
    / "05-评测结果/S23_Experiment1_VFM_Benchmark/metrics/summary_metrics.csv"
)
PHENOTYPE_XLSX = PAPER_ROOT / "植株数据.xlsx"


SEGMENTATION_HEADER = [
    "方法",
    "方法组",
    "类型",
    "训练口径",
    "权重来源",
    "文本提示",
    "时序支持",
    "农业适配",
    "主表优先级",
    "样本范围",
    "序列数",
    "训练序列数",
    "测试序列数",
    "标注帧数",
    "F1",
    "mIoU",
    "HD95像素",
    "边界F分数",
    "时序IoU",
    "面积变异系数",
    "每帧连通域数",
    "推理时间ms",
    "峰值显存GB",
    "外部非黑比例",
    "泄漏能量",
    "失败原因",
    "备注",
]

RECON_HEADER = [
    "方法",
    "方法组",
    "训练口径",
    "权重来源",
    "样本名",
    "代表条件",
    "是否进入下游重建",
    "PSNR_fg",
    "SSIM_fg",
    "LPIPS_fg",
    "外部非黑比例",
    "泄漏能量",
    "高斯数量",
    "网格连通域数",
    "最大连通域比例",
    "株高MAE",
    "冠幅MAE",
    "叶宽MAE",
    "训练时间秒",
    "网格化时间秒",
    "失败原因",
    "备注",
]

EXP5_HEADER = [
    "掩膜来源",
    "样本名",
    "代表条件",
    "人工株高",
    "虚拟株高",
    "株高MAE",
    "株高偏差",
    "人工冠幅",
    "虚拟冠幅",
    "冠幅MAE",
    "冠幅偏差",
    "人工叶长",
    "虚拟叶长",
    "叶长MAE",
    "叶长偏差",
    "人工叶宽",
    "虚拟叶宽",
    "叶宽MAE",
    "叶宽偏差",
    "组内R2",
    "组内MAPE",
    "HD95像素",
    "边界F分数",
    "外部非黑比例",
    "泄漏能量",
    "备注",
]

TRAIT_SUMMARY_HEADER = [
    "性状",
    "n",
    "人工均值",
    "虚拟均值",
    "MAE",
    "RMSE",
    "MAPE百分比",
    "偏差",
    "R2",
]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, header: list[str], rows: Iterable[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in header})


def method_row(rows: list[dict[str, str]], method: str) -> dict[str, str]:
    for row in rows:
        if row.get("method") == method:
            return row
    raise KeyError(f"method {method!r} not found")


def optional_method_row(rows: list[dict[str, str]], method: str) -> dict[str, str] | None:
    for row in rows:
        if row.get("method") == method:
            return row
    return None


def fmt_float(value: object, digits: int = 4) -> str:
    if value is None or value == "":
        return ""
    number = float(value)
    return f"{number:.{digits}f}"


def fmt_hd(value: object) -> str:
    if value is None or value == "":
        return ""
    return f"{float(value):.2f}"


def segmentation_metric_row(
    *,
    name: str,
    group: str,
    model_type: str,
    protocol: str,
    weights: str,
    text_prompt: str,
    temporal: str,
    agri: str,
    priority: str,
    sample_scope: str,
    row: dict[str, str],
    note: str,
    sequence_count: str = "1",
    train_sequence_count: str = "0",
    test_sequence_count: str = "1",
) -> dict[str, str]:
    return {
        "方法": name,
        "方法组": group,
        "类型": model_type,
        "训练口径": protocol,
        "权重来源": weights,
        "文本提示": text_prompt,
        "时序支持": temporal,
        "农业适配": agri,
        "主表优先级": priority,
        "样本范围": sample_scope,
        "序列数": sequence_count,
        "训练序列数": train_sequence_count,
        "测试序列数": test_sequence_count,
        "标注帧数": row.get("eval_frames", row.get("gt_frames", "")),
        "F1": fmt_float(row.get("f1")),
        "mIoU": fmt_float(row.get("miou")),
        "HD95像素": fmt_hd(row.get("hd95_px")),
        "边界F分数": fmt_float(row.get("boundary_f1")),
        "时序IoU": fmt_float(row.get("temporal_iou")),
        "面积变异系数": fmt_float(row.get("area_cv")),
        "每帧连通域数": fmt_float(row.get("component_count_mean")),
        "外部非黑比例": fmt_float(row.get("outside_nonblack_ratio"), 6),
        "泄漏能量": fmt_float(row.get("leakage_energy"), 6),
        "失败原因": "",
        "备注": note,
    }


def pending_segmentation_row(
    name: str,
    group: str,
    model_type: str,
    protocol: str,
    weights: str,
    text_prompt: str,
    temporal: str,
    agri: str,
    priority: str,
    note: str,
) -> dict[str, str]:
    return {
        "方法": name,
        "方法组": group,
        "类型": model_type,
        "训练口径": protocol,
        "权重来源": weights,
        "文本提示": text_prompt,
        "时序支持": temporal,
        "农业适配": agri,
        "主表优先级": priority,
        "样本范围": "待同一GT子集新推理/训练",
        "序列数": "待定",
        "训练序列数": "待定" if "fine-tune" in protocol else "0",
        "测试序列数": "待定",
        "标注帧数": "待定",
        "失败原因": "待新推理" if "fine-tune" not in protocol else "待训练与交叉验证",
        "备注": note,
    }


def blocked_segmentation_row(
    name: str,
    group: str,
    model_type: str,
    protocol: str,
    weights: str,
    text_prompt: str,
    temporal: str,
    agri: str,
    priority: str,
    failure: str,
    note: str,
) -> dict[str, str]:
    row = pending_segmentation_row(
        name,
        group,
        model_type,
        protocol,
        weights,
        text_prompt,
        temporal,
        agri,
        priority,
        note,
    )
    row["样本范围"] = "本轮暂停，不进入实验一当前闭环"
    row["序列数"] = "0"
    row["测试序列数"] = "0"
    row["标注帧数"] = "不适用"
    row["失败原因"] = failure
    return row


def build_experiment1_segmentation_rows() -> list[dict[str, str]]:
    kq = read_csv_rows(KQ_SUMMARY)
    xkl = read_csv_rows(XKL_SUMMARY)
    s23 = read_csv_rows(S23_SUMMARY) if S23_SUMMARY.exists() else []

    def s23_metric_row(
        *,
        method_key: str,
        name: str,
        group: str,
        model_type: str,
        protocol: str,
        weights: str,
        text_prompt: str,
        temporal: str,
        agri: str,
        priority: str,
        note: str,
        train_sequence_count: str = "0",
        test_sequence_count: str = "2",
    ) -> dict[str, str] | None:
        row = optional_method_row(s23, method_key)
        if row is None:
            return None
        asset_note = (
            f" 图件资产见05-图件与论文映射/实验一_视觉基础模型横向对比/"
            f"figure_asset_index.csv，method={method_key}；overlay/error_map/source_data已生成。"
        )
        return segmentation_metric_row(
            name=name,
            group=group,
            model_type=model_type,
            protocol=protocol,
            weights=weights,
            text_prompt=text_prompt,
            temporal=temporal,
            agri=agri,
            priority=priority,
            sample_scope="S23统一GT子集：KongQueZhuYu GT5 + XianKeLai1 GT1",
            row=row,
            note=f"{note}{asset_note}",
            sequence_count="2",
            train_sequence_count=train_sequence_count,
            test_sequence_count=test_sequence_count,
        )

    rows: list[dict[str, str]] = [
        {
            "方法": "SEEM",
            "方法组": "外部基础视觉模型",
            "类型": "可提示/通用交互式分割",
            "训练口径": "zero-shot或交互提示",
            "权重来源": "官方预训练",
            "文本提示": "是",
            "时序支持": "弱",
            "农业适配": "否",
            "主表优先级": "主表",
            "样本范围": "历史人工GT汇总",
            "序列数": "待核原始CSV",
            "训练序列数": "0",
            "测试序列数": "待核原始CSV",
            "标注帧数": "历史汇总",
            "F1": "0.9510",
            "mIoU": "0.9410",
            "HD95像素": "281.90",
            "失败原因": "原始逐帧CSV未在本轮定位",
            "备注": "历史汇总证据来自论文现有证据提取；需补原始逐帧CSV后再作为最终表。",
        },
        blocked_segmentation_row(
            "X-Decoder",
            "外部基础视觉模型",
            "pixel-image-language统一解码",
            "zero-shot文本提示",
            "官方预训练",
            "是",
            "否",
            "否",
            "暂停；不进入当前闭环",
            "用户确认本轮不跑X-Decoder；不进入当前分割主表、图件资产或下游闭环。",
            "保留为文献背景/未来扩展候选；仓库虽已克隆，但不再作为阶段十待办。",
        ),
        blocked_segmentation_row(
            "OpenSeeD",
            "外部基础视觉模型",
            "开放词汇分割与检测",
            "zero-shot文本提示",
            "官方预训练",
            "是",
            "否",
            "否",
            "暂停；不进入当前闭环",
            "用户确认本轮不跑OpenSeeD；不进入当前分割主表、图件资产或下游闭环。",
            "保留为文献背景/未来扩展候选；仓库虽已克隆，但不再作为阶段十待办。",
        ),
        s23_metric_row(
            method_key="Florence2_RES_P2",
            name="Florence-2",
            group="外部基础视觉模型",
            model_type="prompt-based referring-expression segmentation",
            protocol="zero-shot文本指代表达分割P2",
            weights="microsoft/Florence-2-base-ft",
            text_prompt="是",
            temporal="否，本轮逐帧图像模式",
            agri="否",
            priority="主表",
            note="S23同一6帧GT子集实跑；使用REFERRING_EXPRESSION_SEGMENTATION输出polygon并直接栅格化，未加人工后处理。",
        )
        or pending_segmentation_row(
            "Florence-2",
            "外部基础视觉模型",
            "prompt-based视觉基础模型",
            "zero-shot文本提示",
            "官方预训练或HuggingFace",
            "是",
            "弱",
            "否",
            "主表",
            "统一prompt-based grounding/segmentation基线。",
        ),
        s23_metric_row(
            method_key="CLIPSeg_P2",
            name="CLIPSeg",
            group="外部基础视觉模型",
            model_type="文本或图像提示二值分割",
            protocol="zero-shot文本提示P2",
            weights="CIDAS/clipseg-rd64-refined",
            text_prompt="是",
            temporal="否",
            agri="否",
            priority="主表",
            note="S23同一6帧GT子集实跑；阈值0.5，提示词为entire plant excluding pot。",
        )
        or pending_segmentation_row(
            "CLIPSeg",
            "外部基础视觉模型",
            "文本或图像提示二值分割",
            "zero-shot文本提示",
            "官方预训练",
            "是",
            "否",
            "否",
            "主表",
            "轻量CLIP系prompt segmentation基线。",
        ),
        s23_metric_row(
            method_key="SAM2_oracle_box",
            name="SAM2",
            group="外部基础视觉模型",
            model_type="图像/视频对象分割",
            protocol="oracle GT框提示",
            weights="facebook/sam2.1-hiera-large",
            text_prompt="GT框",
            temporal="否，本轮逐帧图像模式",
            agri="否",
            priority="主表上界参照",
            note="S23同一6帧GT子集实跑；使用人工GT bbox作为oracle框提示，不能作为zero-shot文本能力。",
        )
        or pending_segmentation_row(
            "SAM2",
            "外部基础视觉模型",
            "图像/视频对象分割",
            "zero-shot点框提示",
            "官方预训练",
            "点或框",
            "是",
            "否",
            "主表",
            "视频对象传播基线；需记录点/框来源。",
        ),
        s23_metric_row(
            method_key="SAM3_P2",
            name="SAM3单提示词",
            group="外部基础视觉模型",
            model_type="概念或文本分割",
            protocol="zero-shot单提示词P2",
            weights="官方预训练+项目内推理入口",
            text_prompt="是",
            temporal="是",
            agri="否",
            priority="主表",
            note="S23统一6帧GT子集；KongQueZhuYu取实验三A0，XianKeLai1取P2_candidate。",
        )
        or segmentation_metric_row(
            name="SAM3单提示词",
            group="外部基础视觉模型",
            model_type="概念或文本分割",
            protocol="zero-shot单提示词P2",
            weights="官方预训练+项目内推理入口",
            text_prompt="是",
            temporal="是",
            agri="否",
            priority="主表",
            sample_scope="KongQueZhuYu GT5",
            row=method_row(kq, "A0"),
            note="实验三A0；作为RAP-FSAM3-v2的基础模型入口基线。若S23缺失则暂用单样本GT5旧口径。",
        ),
        s23_metric_row(
            method_key="GroundedSAM1_Plant",
            name="Grounded-SAM",
            group="检测加分割",
            model_type="Grounding DINO开放词汇检测框+SAM1",
            protocol="zero-shot文本检测框+逐帧SAM1分割",
            weights="IDEA-Research/grounding-dino-base + sam_vit_h_4b8939",
            text_prompt="plant.",
            temporal="否，本轮逐帧图像模式",
            agri="否",
            priority="主表",
            note="S23同一6帧GT子集实跑；阈值box/text=0.2，取所有plant检测框的SAM1 ViT-H掩膜并集。",
        )
        or pending_segmentation_row(
            "Grounded-SAM",
            "检测加分割",
            "开放词汇检测框+SAM",
            "zero-shot检测框+分割",
            "官方预训练",
            "是",
            "弱",
            "否",
            "主表",
            "需固定plant/potted plant without pot/leaves and stems提示词。",
        ),
        s23_metric_row(
            method_key="GroundedSAM2_Plant",
            name="Grounded-SAM2",
            group="检测加分割",
            model_type="Grounding DINO开放词汇检测框+SAM2",
            protocol="zero-shot文本检测框+逐帧SAM2分割",
            weights="IDEA-Research/grounding-dino-base + facebook/sam2.1-hiera-large",
            text_prompt="plant.",
            temporal="否，本轮逐帧图像模式",
            agri="否",
            priority="主表",
            note="S23同一6帧GT子集实跑；阈值box/text=0.2，取所有plant检测框的SAM2掩膜并集。",
        )
        or pending_segmentation_row(
            "Grounded-SAM2",
            "检测加分割",
            "开放词汇检测框+SAM2",
            "zero-shot检测框+视频传播",
            "官方预训练",
            "是",
            "是",
            "否",
            "可选主表或附表",
            "检查视频传播是否降低逐帧提示不稳定。",
        ),
        s23_metric_row(
            method_key="RAP-FSAM3-v2",
            name="RAP-FSAM3-v2",
            group="本文最终方法",
            model_type="重建感知提示式SAM3派生方法",
            protocol="zero-shot重建感知提示",
            weights="项目内脚本",
            text_prompt="是",
            temporal="是",
            agri="是",
            priority="主表",
            note="S23统一6帧GT子集；KongQueZhuYu取A5c完整RAP-FSAM3-v2，XianKeLai1取A5c语义门控几何修正。",
        )
        or segmentation_metric_row(
            name="RAP-FSAM3-v2",
            group="本文最终方法",
            model_type="重建感知提示式SAM3派生方法",
            protocol="zero-shot重建感知提示",
            weights="项目内脚本",
            text_prompt="是",
            temporal="是",
            agri="是",
            priority="主表",
            sample_scope="KongQueZhuYu GT5",
            row=method_row(kq, "A5c"),
            note="实验三A5c完整RAP-FSAM3-v2；若S23缺失则暂用单样本GT5旧口径。",
        ),
        s23_metric_row(
            method_key="UNet_fewshot_seqcv",
            name="U-Net",
            group="少量监督参照",
            model_type="轻量encoder-decoder分割",
            protocol="few-shot sequence-level 2-fold fine-tune",
            weights="项目内训练；随机初始化；无外部预训练",
            text_prompt="否",
            temporal="否",
            agri="需训练",
            priority="主表参照",
            note="S23同一6帧GT子集实跑；KongQueZhuYu GT5 与 XianKeLai1 GT1 做两折序列级交叉验证，训练/测试序列互斥。",
            train_sequence_count="1/fold",
            test_sequence_count="1/fold",
        )
        or pending_segmentation_row(
            "U-Net",
            "少量监督参照",
            "轻量encoder-decoder分割",
            "few-shot fine-tune",
            "项目内训练",
            "否",
            "否",
            "需训练",
            "主表参照",
            "小样本二分类前景监督基线。",
        ),
        s23_metric_row(
            method_key="DeepLabV3PlusLite_fewshot_seqcv",
            name="DeepLabv3+ lite",
            group="少量监督参照",
            model_type="CNN语义分割",
            protocol="few-shot sequence-level 2-fold fine-tune",
            weights="项目内训练；随机初始化；无外部预训练",
            text_prompt="否",
            temporal="否",
            agri="需训练",
            priority="主表参照",
            note="S23同一6帧GT子集实跑；因当前 torchvision C++ 扩展不可导入，本轮采用项目内轻量 ASPP encoder-decoder 作为 DeepLabv3+ 参照而非官方预训练骨干。",
            train_sequence_count="1/fold",
            test_sequence_count="1/fold",
        )
        or pending_segmentation_row(
            "DeepLabv3+",
            "少量监督参照",
            "CNN语义分割",
            "few-shot fine-tune",
            "MMSegmentation或项目内训练",
            "否",
            "否",
            "需训练",
            "主表参照",
            "经典强监督语义分割基线。",
        ),
        pending_segmentation_row(
            "SegFormer",
            "少量监督参照",
            "Transformer语义分割",
            "few-shot fine-tune",
            "MMSegmentation或HuggingFace",
            "否",
            "否",
            "需训练",
            "主表参照或附表",
            "现代高效Transformer语义分割参照。",
        ),
        pending_segmentation_row(
            "Mask2Former",
            "附表扩展",
            "通用图像分割",
            "类别映射或few-shot fine-tune",
            "官方预训练或MMSegmentation",
            "否",
            "否",
            "需训练或需类别映射",
            "附表优先",
            "强通用分割参照，训练公平性较复杂。",
        ),
    ]
    return rows


def recon_status_row(
    name: str,
    group: str,
    protocol: str,
    weights: str,
    status: str,
    failure: str,
    note: str,
    sample: str = "6个代表性样本",
    condition: str = "实验一下游横向",
) -> dict[str, str]:
    return {
        "方法": name,
        "方法组": group,
        "训练口径": protocol,
        "权重来源": weights,
        "样本名": sample,
        "代表条件": condition,
        "是否进入下游重建": status,
        "失败原因": failure,
        "备注": note,
    }


def build_experiment1_recon_rows() -> list[dict[str, str]]:
    return [
        recon_status_row(
            "SEEM",
            "外部基础视觉模型",
            "zero-shot或交互提示",
            "官方预训练",
            "否，待新重建",
            "SEEM掩膜尚未进入统一2DGS横向流程",
            "已有历史分割汇总；下游重建需按6代表样本重新跑。",
        ),
        recon_status_row(
            "X-Decoder",
            "外部基础视觉模型",
            "zero-shot文本提示",
            "官方预训练",
            "否，本轮暂停",
            "用户确认本轮不跑X-Decoder；不进入当前统一2DGS小闭环",
            "保留为未来扩展候选；当前优先闭环监督参照和已完成S23方法。",
        ),
        recon_status_row(
            "OpenSeeD",
            "外部基础视觉模型",
            "zero-shot文本提示",
            "官方预训练",
            "否，本轮暂停",
            "用户确认本轮不跑OpenSeeD；不进入当前统一2DGS小闭环",
            "保留为未来扩展候选；当前优先闭环监督参照和已完成S23方法。",
        ),
        recon_status_row(
            "Florence-2",
            "外部基础视觉模型",
            "zero-shot文本指代表达分割P2",
            "microsoft/Florence-2-base-ft",
            "否，首轮暂不进入",
            "已完成S23分割；不属于首轮3方法小闭环",
            "S23已有6帧polygon栅格化掩膜；下游可在首轮闭环后扩展。",
        ),
        recon_status_row(
            "CLIPSeg",
            "外部基础视觉模型",
            "zero-shot文本提示",
            "CIDAS/clipseg-rd64-refined",
            "备选小闭环",
            "已完成S23分割；作为Grounded-SAM失败时的替代第三方法",
            "S23已有6帧阈值0.5文本提示掩膜；若需要更轻量的开放词汇参照，可替代Grounded-SAM进入首轮下游。",
        ),
        recon_status_row(
            "SAM2",
            "外部基础视觉模型",
            "oracle GT框提示",
            "facebook/sam2.1-hiera-large",
            "否，首轮暂不进入",
            "已完成S23分割；oracle框不适合作为首轮传导闭环代表",
            "S23已有6帧oracle GT框逐帧SAM2掩膜；这是上界参照，不是zero-shot文本能力。",
        ),
        recon_status_row(
            "SAM3单提示词",
            "外部基础视觉模型",
            "zero-shot单提示词P2",
            "官方预训练+项目内推理入口",
            "优先进入小闭环",
            "已完成S23分割；待统一2DGS/网格/表型流程",
            "A0/P2掩膜已按S23统一6帧重新评估；作为基础模型入口基线进入首轮下游。",
        ),
        recon_status_row(
            "Grounded-SAM",
            "检测加分割",
            "zero-shot检测框+分割",
            "Grounding DINO base + SAM ViT-H",
            "优先进入小闭环",
            "已完成S23分割；待统一2DGS/网格/表型流程",
            "S23已有6帧Grounding DINO plant框 + SAM1 ViT-H掩膜并集；作为开放词汇检测分割代表进入首轮下游。",
        ),
        recon_status_row(
            "Grounded-SAM2",
            "检测加分割",
            "zero-shot检测框+逐帧SAM2分割",
            "Grounding DINO base + SAM2.1 Hiera-L",
            "否，首轮暂不进入",
            "已完成S23分割；首轮优先Grounded-SAM1或CLIPSeg",
            "S23已有6帧Grounding DINO plant框 + SAM2掩膜并集；未使用视频传播。",
        ),
        recon_status_row(
            "RAP-FSAM3-v2",
            "本文最终方法",
            "zero-shot重建感知提示",
            "项目内脚本",
            "优先进入小闭环",
            "已完成S23分割；待按实验一口径统一2DGS/网格/表型流程",
            "S23已有6帧统一分割评估；实验四下游结果不能直接混作实验一下游横向，需另起同口径小闭环。",
        ),
        recon_status_row(
            "U-Net",
            "少量监督参照",
            "few-shot sequence-level 2-fold fine-tune",
            "项目内训练；随机初始化；无外部预训练",
            "否，首轮暂不进入",
            "已完成S23分割；不属于首轮3方法小闭环",
            "U-Net 少量监督参照已完成 S23 两折序列级分割评估；首轮下游仍优先 RAP-FSAM3-v2、SAM3 单提示词、Grounded-SAM 或 CLIPSeg。",
        ),
        recon_status_row(
            "DeepLabv3+ lite",
            "少量监督参照",
            "few-shot sequence-level 2-fold fine-tune",
            "项目内训练；随机初始化；无外部预训练",
            "否，首轮暂不进入",
            "已完成S23分割；不属于首轮3方法小闭环",
            "当前 torchvision C++ 扩展不可导入，故采用项目内轻量 ASPP encoder-decoder 作为 DeepLabv3+ 参照；已完成 S23 两折序列级分割评估。",
        ),
        recon_status_row(
            "SegFormer",
            "少量监督参照",
            "few-shot fine-tune",
            "MMSegmentation或HuggingFace",
            "第三优先级",
            "待训练与交叉验证",
            "SegFormer作为监督参照第三优先级，资源允许再补。",
        ),
    ]


def load_phenotype_pairs() -> dict[str, list[tuple[float, float]]]:
    wb = load_workbook(PHENOTYPE_XLSX, data_only=True, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    pairs = {
        "株高": [],
        "冠幅": [],
        "叶长": [],
        "叶宽": [],
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        pairs["株高"].append((float(row[idx["株高真值"]]), float(row[idx["株高虚拟植"]])))
        pairs["冠幅"].append((float(row[idx["冠幅真值"]]), float(row[idx["冠幅虚拟植"]])))
        for i in (1, 2, 3):
            pairs["叶长"].append((float(row[idx[f"叶长真值{i}"]]), float(row[idx[f"叶长虚拟植{i}"]])))
            pairs["叶宽"].append((float(row[idx[f"叶宽真值{i}"]]), float(row[idx[f"叶宽虚拟植{i}"]])))
    return pairs


def trait_metrics(values: list[tuple[float, float]]) -> dict[str, float]:
    n = len(values)
    manual = [x for x, _ in values]
    virtual = [y for _, y in values]
    errors = [y - x for x, y in values]
    abs_errors = [abs(e) for e in errors]
    sq_errors = [e * e for e in errors]
    ape = [abs(y - x) / x * 100.0 for x, y in values if x != 0]
    manual_mean = sum(manual) / n
    ss_res = sum((x - y) ** 2 for x, y in values)
    ss_tot = sum((x - manual_mean) ** 2 for x in manual)
    r2 = 1.0 - ss_res / ss_tot if ss_tot else math.nan
    return {
        "n": n,
        "manual_mean": manual_mean,
        "virtual_mean": sum(virtual) / n,
        "mae": sum(abs_errors) / n,
        "rmse": math.sqrt(sum(sq_errors) / n),
        "mape": sum(ape) / len(ape),
        "bias": sum(errors) / n,
        "r2": r2,
    }


def build_trait_summary_rows() -> tuple[list[dict[str, str]], dict[str, dict[str, float]]]:
    pairs = load_phenotype_pairs()
    metrics = {trait: trait_metrics(values) for trait, values in pairs.items()}
    rows = []
    for trait in ["株高", "冠幅", "叶长", "叶宽"]:
        m = metrics[trait]
        rows.append(
            {
                "性状": trait,
                "n": str(int(m["n"])),
                "人工均值": fmt_float(m["manual_mean"]),
                "虚拟均值": fmt_float(m["virtual_mean"]),
                "MAE": fmt_float(m["mae"]),
                "RMSE": fmt_float(m["rmse"]),
                "MAPE百分比": fmt_float(m["mape"]),
                "偏差": fmt_float(m["bias"]),
                "R2": fmt_float(m["r2"]),
            }
        )
    return rows, metrics


def build_experiment5_rows(metrics: dict[str, dict[str, float]]) -> list[dict[str, str]]:
    kq = method_row(read_csv_rows(KQ_SUMMARY), "A5c")
    xkl = method_row(read_csv_rows(XKL_SUMMARY), "A5c")

    def metric_text(key: str) -> str:
        return ";".join(f"{trait}={metrics[trait][key]:.4f}" for trait in ["株高", "冠幅", "叶长", "叶宽"])

    main = {
        "掩膜来源": "RAP-FSAM3-v2/ForeSplat完整流程",
        "样本名": "21株10品种汇总",
        "代表条件": "人工-虚拟表型验证；非多掩膜来源横向",
        "人工株高": fmt_float(metrics["株高"]["manual_mean"]),
        "虚拟株高": fmt_float(metrics["株高"]["virtual_mean"]),
        "株高MAE": fmt_float(metrics["株高"]["mae"]),
        "株高偏差": fmt_float(metrics["株高"]["bias"]),
        "人工冠幅": fmt_float(metrics["冠幅"]["manual_mean"]),
        "虚拟冠幅": fmt_float(metrics["冠幅"]["virtual_mean"]),
        "冠幅MAE": fmt_float(metrics["冠幅"]["mae"]),
        "冠幅偏差": fmt_float(metrics["冠幅"]["bias"]),
        "人工叶长": fmt_float(metrics["叶长"]["manual_mean"]),
        "虚拟叶长": fmt_float(metrics["叶长"]["virtual_mean"]),
        "叶长MAE": fmt_float(metrics["叶长"]["mae"]),
        "叶长偏差": fmt_float(metrics["叶长"]["bias"]),
        "人工叶宽": fmt_float(metrics["叶宽"]["manual_mean"]),
        "虚拟叶宽": fmt_float(metrics["叶宽"]["virtual_mean"]),
        "叶宽MAE": fmt_float(metrics["叶宽"]["mae"]),
        "叶宽偏差": fmt_float(metrics["叶宽"]["bias"]),
        "组内R2": metric_text("r2"),
        "组内MAPE": metric_text("mape"),
        "HD95像素": f"KongQueZhuYu_GT5={fmt_hd(kq['hd95_px'])};XianKeLai1_GT1={fmt_hd(xkl['hd95_px'])}",
        "边界F分数": f"KongQueZhuYu_GT5={fmt_float(kq['boundary_f1'])};XianKeLai1_GT1={fmt_float(xkl['boundary_f1'])}",
        "外部非黑比例": f"KongQueZhuYu_GT5={fmt_float(kq['outside_nonblack_ratio'], 6)};XianKeLai1_GT1={fmt_float(xkl['outside_nonblack_ratio'], 6)}",
        "泄漏能量": f"KongQueZhuYu_GT5={fmt_float(kq['leakage_energy'], 6)};XianKeLai1_GT1={fmt_float(xkl['leakage_energy'], 6)}",
        "备注": "来自植株数据.xlsx；叶宽MAPE最高，支持边界敏感性判断；SEEM/SAM3/FSAM3等多掩膜来源表型重建尚未完成。",
    }

    pending = []
    for source, note in [
        ("SEEM", "已有历史分割汇总，但未完成SEEM掩膜驱动的同口径2DGS和表型测量。"),
        ("SAM3单提示词", "已有GT5/GT1分割结果，但未完成单提示词掩膜驱动的同口径2DGS和表型测量。"),
        ("FSAM3基线", "归入实验三内部方法族；若要做表型敏感性，需补统一下游重建。"),
        ("人工腐蚀/膨胀掩膜", "控制变量敏感性分析待生成扰动掩膜并重建/测量。"),
    ]:
        pending.append(
            {
                "掩膜来源": source,
                "样本名": "待6代表样本统一重建",
                "代表条件": "多掩膜来源表型敏感性",
                "备注": f"待新重建：{note}",
            }
        )

    return [main, *pending]


def main() -> None:
    write_csv(EXP1_SEGMENTATION_TABLE, SEGMENTATION_HEADER, build_experiment1_segmentation_rows())
    write_csv(EXP1_RECON_TABLE, RECON_HEADER, build_experiment1_recon_rows())
    trait_rows, trait_metric_map = build_trait_summary_rows()
    write_csv(EXP5_TRAIT_SUMMARY, TRAIT_SUMMARY_HEADER, trait_rows)
    write_csv(EXP5_TABLE, EXP5_HEADER, build_experiment5_rows(trait_metric_map))

    print(f"Wrote {EXP1_SEGMENTATION_TABLE}")
    print(f"Wrote {EXP1_RECON_TABLE}")
    print(f"Wrote {EXP5_TRAIT_SUMMARY}")
    print(f"Wrote {EXP5_TABLE}")


if __name__ == "__main__":
    main()
